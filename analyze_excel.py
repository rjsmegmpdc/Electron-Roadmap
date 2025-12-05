#!/usr/bin/env python
import openpyxl
import json

excel_file = r'C:\Users\smhar\Downloads\98047 Modern Workspace modernization Financial Tracker.xlsx'

try:
    wb = openpyxl.load_workbook(excel_file)
    
    print("=" * 80)
    print("EXCEL FILE STRUCTURE ANALYSIS")
    print("=" * 80)
    print(f"\nSheet Names: {wb.sheetnames}\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"SHEET: {sheet_name}")
        print(f"{'='*80}")
        print(f"Max Rows: {ws.max_row}, Max Columns: {ws.max_column}\n")
        
        # Get first row (headers)
        headers = []
        for col in range(1, min(ws.max_column + 1, 20)):  # First 20 columns
            cell = ws.cell(row=1, column=col)
            headers.append(cell.value)
        
        print(f"Headers (first 20 columns):")
        for i, h in enumerate(headers, 1):
            print(f"  {i}. {h}")
        
        # Show first 5 data rows
        print(f"\nFirst 5 Data Rows:")
        for row in range(2, min(7, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(ws.max_column + 1, 20)):
                cell = ws.cell(row=row, column=col)
                row_data.append(cell.value)
            print(f"  Row {row}: {row_data[:10]}...")  # Show first 10 cols
        
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
